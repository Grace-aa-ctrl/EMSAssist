from datasets import load_dataset
from transformers import AutoTokenizer
from transformers import AutoModelForSequenceClassification, AutoModel
from transformers import TrainingArguments
import argparse
from datetime import datetime
import os
import sys
sys.path.append("/home/liuyi/transformers")
from collections import defaultdict
import utils.file_utils as util
from utils.nemsis_processor import NEMSIS_Processor as NP
from utils.medication_processor import Medication_Processor as MP
from utils.vitals_processor import Vital_Processor as VP
import random
import numpy as np
import torch
from torch import nn
# from torch.utils.data import DataLoader
from torch.optim import AdamW
from torch.nn import CrossEntropyLoss, MSELoss
from transformers import get_scheduler
from tqdm.auto import tqdm
import evaluate
from torch.nn.utils.rnn import pad_sequence, pack_padded_sequence, pad_packed_sequence
from torch.utils.data import DataLoader, Dataset, TensorDataset, ConcatDataset
from torch.utils.tensorboard import SummaryWriter
# from torchsummary import summary
# from torchviz import make_dot

# import multiprocessing as mp
import fcntl
import pandas as pd
import openpyxl
from openpyxl import Workbook
import re
import utils.file_utils as util

import utils.training_utils as tu

def set_seed(seed):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True

def compute_topk_accuracy(logits, labels, k):
    _, topk_indices = torch.topk(logits, k, dim=-1)
    topk_correct = topk_indices.eq(labels.view(-1, 1).expand_as(topk_indices))
    topk_accuracy = topk_correct.sum().item() / labels.size(0)
    return topk_accuracy

# def compute_topk_accuracy(logits, labels, k):
#     # Get the indices of the top-k predictions
#     _, topk_indices = torch.topk(logits, k, dim=-1)
    
#     # Convert one-hot encoded labels to indices of true classes
#     true_class_indices = labels.nonzero(as_tuple=True)[1].view(labels.size(0), -1)
    
#     # Initialize the count of correct predictions
#     topk_correct = 0

#     # Check if any of the top-k predictions are in the true class indices
#     for i in range(labels.size(0)):
#         if any(idx in topk_indices[i] for idx in true_class_indices[i]):
#             topk_correct += 1

#     # Compute top-k accuracy
#     topk_accuracy = topk_correct / labels.size(0)
#     return topk_accuracy


def save_results_to_csv(results_dict, filename):
    df = pd.DataFrame(results_dict).T  # Convert dict of dicts to DataFrame
    with open(filename, 'a') as f:
        fcntl.flock(f, fcntl.LOCK_EX)
        df.to_csv(f, header=f.tell()==0, index=False)
        fcntl.flock(f, fcntl.LOCK_UN)

def get_protocol_cell_indices(results_dict, args, row_idx, col_width_each_metric = 4):
    col_base_idx = 0
    if args.do_med_type and args.do_quantity:
        col_base_idx += 3
    elif args.do_quantity:
        col_base_idx += 2
    elif args.do_med_type:
        col_base_idx += 1

    col_idx = col_base_idx
    results_dict["protocol"]["protocol_top1_accuracy"].extend([(row_idx,col_idx), (row_idx+1,col_idx)])
    results_dict["protocol"]["protocol_top3_accuracy"].extend([(row_idx,col_idx + col_width_each_metric), (row_idx+1,col_idx + col_width_each_metric)])   
    results_dict["protocol"]["protocol_top5_accuracy"].extend([(row_idx,col_idx + 2*col_width_each_metric), (row_idx+1,col_idx + 2*col_width_each_metric)])

def get_med_type_cell_indices(results_dict, args, row_idx, col_width_each_metric = 4):
    col_base_idx = 0
    if args.do_protocol and args.do_quantity:
        col_base_idx += 3
    elif args.do_quantity:
        col_base_idx += 2
    elif args.do_protocol:
        col_base_idx += 1

    col_idx = col_base_idx
    results_dict["med_type"]["med_type_top1_accuracy"].extend([(row_idx,col_idx), (row_idx+1,col_idx)])
    results_dict["med_type"]["med_type_top3_accuracy"].extend([(row_idx,col_idx + col_width_each_metric), (row_idx+1,col_idx + col_width_each_metric)])   
    results_dict["med_type"]["med_type_top5_accuracy"].extend([(row_idx,col_idx + 2*col_width_each_metric), (row_idx+1,col_idx + 2*col_width_each_metric)])

def get_quantity_cell_indices(results_dict, args, row_idx, col_width_each_metric = 4):
    col_base_idx = 0
    if args.do_protocol and args.do_med_type:
        col_base_idx += 3
    elif args.do_med_type:
        col_base_idx += 2
    elif args.do_protocol:
        col_base_idx += 1

    col_idx = col_base_idx
    results_dict["quantity"]["quantity_mse"].extend([(row_idx,col_idx), (row_idx+1,col_idx)])
    results_dict["quantity"]["quantity_pearsonr"].extend([(row_idx,col_idx + col_width_each_metric), (row_idx+1,col_idx + col_width_each_metric)])   
    results_dict["quantity"]["quantity_spearmanr"].extend([(row_idx,col_idx + 2*col_width_each_metric), (row_idx+1,col_idx + 2*col_width_each_metric)])


def get_multi_task_excel_cell_indices(results_dict, args):
    row_idx = (int(args.nemsis_year) - 2019) * 2                        # year defines the row_index
    if args.do_protocol:
        get_protocol_cell_indices(results_dict, args, row_idx)
    if args.do_med_type:
        get_med_type_cell_indices(results_dict, args, row_idx)
    if args.do_quantity:
        get_quantity_cell_indices(results_dict, args, row_idx)


def save_results_to_excel(results_dict, excel_file):
    # # Lock the file
    # if not os.path.exists(excel_file):
    #     append_write = 'w'
    # else:
    #     append_write = 'a+b'

    # with open(excel_file, append_write) as file:
    #     fcntl.flock(file, fcntl.LOCK_EX)

    # Load the workbook or create a new one if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(excel_file)
        # print(f"Loaded existing workbook {file_path}")
    except FileNotFoundError:
        workbook = Workbook()
        # print(f"Created new workbook {file_path}")

    # workbook = openpyxl.load_workbook(excel_file)
    
    # Iterate over the results dictionary
    for sheet_name, metric_data_dict in results_dict.items():
        # Get the worksheet by name or create a new one if it doesn't exist
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
        for metric_name, data in metric_data_dict.items():
            # print("writing to the sheet %s, metric %s, data %s" % (sheet_name, metric_name, data))
            if len(data) == 4:
                validate_idx, test_idx, validate_value, test_value = data[0], data[1], data[2], data[3]
                # print("writing to the excel...")
                # print(validate_idx, test_idx, validate_value, test_value)
                # Write data to specific cells
                # for row_idx, row_data in enumerate(data, start=1):
                #     for col_idx, cell_value in enumerate(row_data, start=1):
                worksheet.cell(row=validate_idx[0] + 1, column=validate_idx[1]+1, value=validate_value)
                worksheet.cell(row=test_idx[0]+1, column=test_idx[1]+1, value=test_value)

    # Save the workbook
    workbook.save(excel_file)

        # Unlock the file
        # fcntl.flock(file, fcntl.LOCK_UN)

    print(f"Results saved to: {excel_file}")

class VitalCollate:
    def __init__(self, pre_padding = False, padding_value = 0.0):
        self.pre_padding = pre_padding
        self.padding_value = padding_value

    def __call__(self, batch):
        return self.collate_fn(batch)

    def collate_fn(self, batch):
        vital_data = [item['vital'] for item in batch]
        lengths = torch.tensor([len(seq) for seq in vital_data])
        max_len = max(lengths)

        padded_vital_data = pad_sequence(vital_data, batch_first=True)
        if self.pre_padding:
            padded_sequences = []
            for seq in vital_data:
                length = seq.size(0)
                padding = torch.full((max_len - length, *seq.size()[1:]), self.padding_value, dtype=seq.dtype, device=seq.device)
                padded_seq = torch.cat((padding, seq), dim=0)
                padded_sequences.append(padded_seq)
            padded_vital_data = torch.stack(padded_sequences)

        non_vital_data = {k: torch.stack([item[k] for item in batch]) for k in batch[0] if k not in ['vital']}
        return {**non_vital_data, 'vital': padded_vital_data, 'lengths': lengths}


def train_emsglass(train_dataloader,
                   eval_dataloader,
                   args,
                   protocol_label_name,
                   med_type_label_name,
                   quantity_label_name,
                   best_model_path,
                   text_model=None,
                   vital_model=None,
                   scene_model=None,    
                   text_vital_model=None,
                   text_scene_model=None,
                   vital_scene_model=None,
                   text_vital_scene_model=None):
    # Mapping models to their expected input configurations
    model_input_map = {
        "text_model": (text_model, ["input_ids", "attention_mask", "token_type_ids"]),
        "vital_model": (vital_model, ["vital"]),
        "scene_model": (scene_model, ["scene"]),
        "text_vital_model": (text_vital_model, ["input_ids", "attention_mask", "token_type_ids", "vital"]),
        "text_scene_model": (text_scene_model, ["input_ids", "attention_mask", "token_type_ids", "scene"]),
        "vital_scene_model": (vital_scene_model, ["vital", "scene"]),
        "text_vital_scene_model": (text_vital_scene_model, ["input_ids", "attention_mask", "token_type_ids", "vital", "scene"])
    }

    # Device setup
    device = torch.device(f"cuda:{args.device}" if torch.cuda.is_available() else "cpu")

    # Metrics and counters setup for each model
    tasks = {
        "protocol": {"do_task": args.do_protocol, "metric": evaluate.load("accuracy"), "topk": [1, 3, 5], "loss_fn": CrossEntropyLoss()},
        "med_type": {"do_task": args.do_med_type, "metric": evaluate.load("accuracy"), "topk": [1, 3, 5], "loss_fn": CrossEntropyLoss()},
        "quantity": {"do_task": args.do_quantity, "metrics": {"mse": evaluate.load("mse"), 
                                                             "pearsonr": evaluate.load("pearsonr"), 
                                                             "spearmanr": evaluate.load("spearmanr")}, "loss_fn": MSELoss()}
    }

    # Train each model
    for model_name, (model, required_inputs) in model_input_map.items():
        if model is None:
            continue

        print(f"[train_emsglass] begins training")
        # print(f"Training {model_name}...")
        optimizer = AdamW(model.parameters(), lr=5e-5)
        num_epochs = args.epoch
        num_training_steps = num_epochs * len(train_dataloader)
        lr_scheduler = get_scheduler(
            name="linear", optimizer=optimizer, num_warmup_steps=0, num_training_steps=num_training_steps
        )
        model.to(device)

        # Early stopping parameters
        patience = args.patience  # Number of epochs to wait for improvement
        best_val_accuracy = float("-inf")
        best_evaluate_results = None
        epochs_without_improvement = 0

        for epoch in range(num_epochs):
            model.train()
            progress_bar = tqdm(range(len(train_dataloader)), desc=f"[Training {model_name} ==> EPOCH {epoch}]")
            # Training loop
            for batch in train_dataloader:
                batch = {k: v.to(device) for k, v in batch.items()}
                # print(f"[debug] batch input in train_emsglass: " , batch)
                
                # Rename one_hot_scene to scene
                if "one_hot_scene" in batch:
                    batch["scene"] = batch.pop("one_hot_scene")

                # Prepare inputs based on model requirements
                inputs = {k: batch[k] for k in required_inputs if k in batch}

                # Forward pass and loss computation
                # optimizer.zero_grad()
                protocol_logits, med_type_logits, quantity_output, *_ = model(**inputs)
                # print(f"[debug] quantity output in train_emsglass: " , quantity_output)

                batch_loss = 0.0
                # Protocol classification task
                if tasks["protocol"]["do_task"]:
                    protocol_labels = batch[protocol_label_name]
                    protocol_loss = tasks["protocol"]["loss_fn"](protocol_logits, protocol_labels)
                    batch_loss += protocol_loss

                # Medication type classification task
                if tasks["med_type"]["do_task"]:
                    med_type_labels = batch[med_type_label_name]
                    med_type_loss = tasks["med_type"]["loss_fn"](med_type_logits, med_type_labels)
                    batch_loss += med_type_loss

                # Quantity regression task
                if tasks["quantity"]["do_task"]:
                    quantity_labels = batch[quantity_label_name]
                    quantity_loss = tasks["quantity"]["loss_fn"](quantity_output.squeeze(-1), quantity_labels)
                    batch_loss += quantity_loss

                batch_loss.backward()
                optimizer.step()
                lr_scheduler.step()
                optimizer.zero_grad()
                progress_bar.update(1)

            progress_bar.close()
            # progress_bar.clear()

            # Forward only the current model to the evaluation function
            eval_kwargs = {name: None for name in model_input_map.keys()}
            eval_kwargs[model_name] = model  # Set the current model
            evaluate_results = evaluate_emsglass(
                eval_dataloader,
                args,
                protocol_label_name,
                med_type_label_name,
                quantity_label_name,
                # choose_user_study_samples=False,
                **eval_kwargs
            )
            print(f"Results after epoch {epoch}:")
            tu.print_evaluate_results(args, evaluate_results)

            if args.do_protocol:
                val_accuracy = evaluate_results[model_name]["protocol"]["top-3"]
            elif args.do_med_type:
                val_accuracy = evaluate_results[model_name]["med_type"]["top-3"]
            else:
                val_accuracy = -evaluate_results[model_name]["quantity"]["mse"]

            # Early stopping check
            if val_accuracy > best_val_accuracy:
                best_val_accuracy = val_accuracy
                best_evaluate_results = evaluate_results
                epochs_without_improvement = 0

                # Save the best model
                torch.save(model.state_dict(), best_model_path)
                print(f"Best {model_name} model saved in {best_model_path} with validation accuracy: {val_accuracy if (args.do_protocol or args.do_med_type) else -val_accuracy}")

            else:
                epochs_without_improvement += 1

            if epochs_without_improvement >= patience or epoch == num_epochs - 1:
                print(f"(Early) stopping after epoch {epoch}")
                # progress_bar.close()
                return best_evaluate_results
            

def evaluate_emsglass(test_dataloader, 
                      args, 
                      protocol_label_name, 
                      med_type_label_name, 
                      quantity_label_name,
                      text_model=None,
                      vital_model=None,
                      scene_model=None,
                      text_vital_model=None,
                      text_scene_model=None,
                      vital_scene_model=None,
                      text_vital_scene_model=None):
    
    # Mapping models to their expected input configurations
    model_input_map = {
        "text_model": (text_model, ["input_ids", "attention_mask", "token_type_ids"]),
        "vital_model": (vital_model, ["vital"]),
        "scene_model": (scene_model, ["scene"]),
        "text_vital_model": (text_vital_model, ["input_ids", "attention_mask", "token_type_ids", "vital"]),
        "text_scene_model": (text_scene_model, ["input_ids", "attention_mask", "token_type_ids", "scene"]),
        "vital_scene_model": (vital_scene_model, ["vital", "scene"]),
        "text_vital_scene_model": (text_vital_scene_model, ["input_ids", "attention_mask", "token_type_ids", "vital", "scene"])
    }

    # Device setup
    device = torch.device(f"cuda:{args.device}" if torch.cuda.is_available() else "cpu")

    # Results dictionary and temporary storage
    all_results = {}

    # Evaluate each model
    for model_name, (model, required_inputs) in model_input_map.items():
        if model is None:
            continue
        # print(f"[evaluate_emsglass] begins evaluation")
        model.to(device)
        model.eval()

        # Metrics and counters setup for each model
        tasks = {
            "protocol": {"do_task": args.do_protocol, "metric": evaluate.load("accuracy"), "topk": [1, 3, 5]},
            "med_type": {"do_task": args.do_med_type, "metric": evaluate.load("accuracy"), "topk": [1, 3, 5]},
            "quantity": {"do_task": args.do_quantity, "metrics": {"mse": evaluate.load("mse"), "pearsonr": evaluate.load("pearsonr"), "spearmanr": evaluate.load("spearmanr")}}
        }
        progress_bar = tqdm(range(len(test_dataloader)), desc=f"Evaluating {model_name}")
        total_samples = 0

        # Evaluation loop for the current model
        for batch in test_dataloader:
            batch = {k: v.to(device) for k, v in batch.items()}
            
            # Rename one_hot_scene to scene
            if "one_hot_scene" in batch:
                batch["scene"] = batch.pop("one_hot_scene")

            # Prepare inputs based on model requirements
            inputs = {k: batch[k] for k in required_inputs if k in batch}
            
            with torch.no_grad():
                # Forward pass
                protocol_logits, med_type_logits, quantity_output, *_ = model(**inputs)
                # print(f"[debug] quantity output in evaluate_emsglass: " , quantity_output)

            # Task evaluation
            if tasks["protocol"]["do_task"]:
                predictions = torch.argmax(protocol_logits, dim=-1)
                label = batch[protocol_label_name]
                tasks["protocol"]["metric"].add_batch(predictions=predictions, references=label)
                for k in tasks["protocol"]["topk"]:
                    accuracy = compute_topk_accuracy(protocol_logits, label, k=k) * len(label)
                    tasks["protocol"].setdefault(f"top{k}", 0)
                    tasks["protocol"][f"top{k}"] += accuracy
                # if model_name in model_predictions:
                #     model_predictions[model_name]["protocol"].append(predictions == label)
                # if choose_user_study_samples and model_name == "text_vital_scene_model":
                #     model_predictions["protocol"].append(predictions == label)

            if tasks["med_type"]["do_task"]:
                predictions = torch.argmax(med_type_logits, dim=-1)
                label = batch[med_type_label_name]
                tasks["med_type"]["metric"].add_batch(predictions=predictions, references=label)
                for k in tasks["med_type"]["topk"]:
                    accuracy = compute_topk_accuracy(med_type_logits, label, k=k) * len(label)
                    tasks["med_type"].setdefault(f"top{k}", 0)
                    tasks["med_type"][f"top{k}"] += accuracy
                # if model_name in model_predictions:
                #     model_predictions[model_name]["med_type"].append(predictions == label)
                # if choose_user_study_samples and model_name == "text_vital_scene_model":
                #     model_predictions["med_type"].append(predictions == label)

            if tasks["quantity"]["do_task"]:
                for name, metric in tasks["quantity"]["metrics"].items():
                    # print(f"[debug] predictions {quantity_output.squeeze(-1)}, reference {batch[quantity_label_name]}")
                    metric.add_batch(predictions=quantity_output.squeeze(-1), references=batch[quantity_label_name])

            total_samples += len(batch["pcr"])
            progress_bar.update(1)

        progress_bar.close()
        # Compute results for the current model
        model_results = {}
        for task_name, task_info in tasks.items():
            if task_info["do_task"]:
                if task_name == "quantity":
                    model_results[task_name] = {name: metric.compute()[name] for name, metric in task_info["metrics"].items()}
                else:
                    accuracy = task_info["metric"].compute()["accuracy"]
                    top_accuracies = {f"top-{k}": task_info[f"top{k}"] / total_samples for k in task_info["topk"]}
                    model_results[task_name] = {"accuracy": accuracy, **top_accuracies}

        all_results[model_name] = model_results

    # # Print all results at the end
    # for model_name, results in all_results.items():
    #     print(f"\nResults for {model_name}:")
    #     for task, metrics in results.items():
    #         print(f"{task} Metrics: {metrics}")

    return all_results        


def evaluate_emsglass_user_study(   test_dataloader, 
                                    args, 
                                    protocol_label_name, 
                                    med_type_label_name, 
                                    quantity_label_name,
                                    text_model=None,
                                    vital_model=None,
                                    scene_model=None,
                                    text_vital_model=None,
                                    text_scene_model=None,
                                    vital_scene_model=None,
                                    text_vital_scene_model=None,
                                    choose_user_study_samples=False,
                                    user_study_file_path=None  ):
    
    # Mapping models to their expected input configurations
    model_input_map = {
        "text_model": (text_model, ["input_ids", "attention_mask", "token_type_ids"]),
        "vital_model": (vital_model, ["vital"]),
        "scene_model": (scene_model, ["scene"]),
        "text_vital_model": (text_vital_model, ["input_ids", "attention_mask", "token_type_ids", "vital"]),
        "text_scene_model": (text_scene_model, ["input_ids", "attention_mask", "token_type_ids", "scene"]),
        "vital_scene_model": (vital_scene_model, ["vital", "scene"]),
        "text_vital_scene_model": (text_vital_scene_model, ["input_ids", "attention_mask", "token_type_ids", "vital", "scene"])
    }

    # Device setup
    device = torch.device(f"cuda:{args.device}" if torch.cuda.is_available() else "cpu")

    # Results dictionary and temporary storage
    all_results = {}
    user_study_samples = []

    # Temporary storage for model predictions
    model_predictions = {
        "text_model": {"protocol": [], "med_type": []},
        "text_vital_model": {"protocol": [], "med_type": []},
        "text_vital_scene_model": {"protocol": [], "med_type": []}
    }

    # Evaluate each model
    for model_name, (model, required_inputs) in model_input_map.items():
        if model is None:
            continue
        model.to(device)
        model.eval()

        # Metrics and counters setup for each model
        tasks = {
            "protocol": {"do_task": args.do_protocol, "metric": evaluate.load("accuracy"), "topk": [1, 3, 5]},
            "med_type": {"do_task": args.do_med_type, "metric": evaluate.load("accuracy"), "topk": [1, 3, 5]},
            "quantity": {"do_task": args.do_quantity, "metrics": {"mse": evaluate.load("mse"), "pearsonr": evaluate.load("pearsonr"), "spearmanr": evaluate.load("spearmanr")}}
        }
        # progress_bar = tqdm(range(len(test_dataloader)), desc=f"Evaluating {model_name}")
        total_samples = 0

        # Evaluation loop for the current model
        for batch in test_dataloader:
            batch = {k: v.to(device) for k, v in batch.items()}
            
            # Rename one_hot_scene to scene
            if "one_hot_scene" in batch:
                batch["scene"] = batch.pop("one_hot_scene")

            # Prepare inputs based on model requirements
            inputs = {k: batch[k] for k in required_inputs if k in batch}
            # if model_name == "text_vital_scene_model":
            #     print("batch inputs: ", inputs)
            #     print("batch inputs input_ids shape: ", inputs["input_ids"].shape)
            #     print("batch inputs vital shape: ", inputs["vital"].shape)
            #     print("batch inputs scene shape: ", inputs["scene"].shape)
            #     return
            with torch.no_grad():
                # Forward pass
                protocol_logits, med_type_logits, quantity_output, *_ = model(**inputs)

            # Task evaluation
            if tasks["protocol"]["do_task"]:
                predictions = torch.argmax(protocol_logits, dim=-1)
                label = batch[protocol_label_name]
                tasks["protocol"]["metric"].add_batch(predictions=predictions, references=label)
                for k in tasks["protocol"]["topk"]:
                    accuracy = compute_topk_accuracy(protocol_logits, label, k=k) * len(label)
                    tasks["protocol"].setdefault(f"top{k}", 0)
                    tasks["protocol"][f"top{k}"] += accuracy
                if model_name in model_predictions:
                    model_predictions[model_name]["protocol"].append(predictions == label)
                # if choose_user_study_samples and model_name == "text_vital_scene_model":
                #     model_predictions["protocol"].append(predictions == label)

            if tasks["med_type"]["do_task"]:
                predictions = torch.argmax(med_type_logits, dim=-1)
                label = batch[med_type_label_name]
                tasks["med_type"]["metric"].add_batch(predictions=predictions, references=label)
                for k in tasks["med_type"]["topk"]:
                    accuracy = compute_topk_accuracy(med_type_logits, label, k=k) * len(label)
                    tasks["med_type"].setdefault(f"top{k}", 0)
                    tasks["med_type"][f"top{k}"] += accuracy
                if model_name in model_predictions:
                    model_predictions[model_name]["med_type"].append(predictions == label)
                # if choose_user_study_samples and model_name == "text_vital_scene_model":
                #     model_predictions["med_type"].append(predictions == label)

            if tasks["quantity"]["do_task"]:
                for name, metric in tasks["quantity"]["metrics"].items():
                    metric.add_batch(predictions=quantity_output.squeeze(-1), references=batch[quantity_label_name])

            total_samples += len(batch["pcr"])
            # progress_bar.update(1)

        if model_name in model_predictions:
            model_predictions[model_name]["protocol"] = torch.cat(model_predictions[model_name]["protocol"])
            model_predictions[model_name]["med_type"] = torch.cat(model_predictions[model_name]["med_type"])

        # # Collect samples that only the specified model gets correct
        # if choose_user_study_samples and model_name == "text_vital_scene_model":
        #     correct_protocol = torch.cat(model_predictions["protocol"])
        #     correct_med_type = torch.cat(model_predictions["med_type"])
        #     for i, (protocol_correct, med_type_correct) in enumerate(zip(correct_protocol, correct_med_type)):
        #         if protocol_correct and med_type_correct:
        #             user_study_samples.append(test_dataloader.dataset[i])

        # Compute results for the current model
        model_results = {}
        for task_name, task_info in tasks.items():
            if task_info["do_task"]:
                if task_name == "quantity":
                    model_results[task_name] = {name: metric.compute()[name] for name, metric in task_info["metrics"].items()}
                else:
                    accuracy = task_info["metric"].compute()["accuracy"]
                    top_accuracies = {f"Top-{k}": task_info[f"top{k}"] / total_samples for k in task_info["topk"]}
                    model_results[task_name] = {"Accuracy": accuracy, **top_accuracies}

        all_results[model_name] = model_results

    # Collect samples based on the new criteria
    if choose_user_study_samples:
        correct_protocol_tvs = model_predictions["text_vital_scene_model"]["protocol"]
        correct_med_type_tvs = model_predictions["text_vital_scene_model"]["med_type"]
        incorrect_protocol_text = ~model_predictions["text_model"]["protocol"]
        incorrect_med_type_text = ~model_predictions["text_model"]["med_type"]
        incorrect_protocol_text_vital = ~model_predictions["text_vital_model"]["protocol"]
        incorrect_med_type_text_vital = ~model_predictions["text_vital_model"]["med_type"]

        for i, (protocol_correct_tvs, med_type_correct_tvs, protocol_incorrect_text, med_type_incorrect_text, protocol_incorrect_text_vital, med_type_incorrect_text_vital) in enumerate(
                zip(correct_protocol_tvs, correct_med_type_tvs, incorrect_protocol_text, incorrect_med_type_text, incorrect_protocol_text_vital, incorrect_med_type_text_vital)):
            if protocol_correct_tvs and med_type_correct_tvs and protocol_incorrect_text and med_type_incorrect_text and protocol_incorrect_text_vital and med_type_incorrect_text_vital:
                user_study_samples.append(test_dataloader.dataset[i])

    # Print all results at the end
    for model_name, results in all_results.items():
        print(f"\nResults for {model_name}:")
        for task, metrics in results.items():
            print(f"{task} Metrics: {metrics}")

    # Return selected samples if requested
    if choose_user_study_samples:
        print(f"\nSelected user study samples {len(user_study_samples)}")
        assert user_study_file_path is not None, "User study file path must be provided"
        sample_pcr = []
        for sample in user_study_samples:
            sample_pcr.append(str(sample['pcr'].item()))
            print(sample['pcr'], type(sample['pcr']))
        sample_user_study_examples(args, sample_pcr, user_study_file_path)

        return all_results, user_study_samples


    return all_results   

def sample_user_study_examples(args, pcr_keys_to_add, user_study_file_path):
    # Setup paths
    data_dir = os.path.join(args.home_dir, args.data_folder)
    cache_dir = os.path.join(data_dir, args.cache_folder)
    my_np = NP(args.nemsis_dir, data_dir, cache_dir, args.nemsis_year)

    # using the validation file for testing
    test_file_scene_path = os.path.join(cache_dir, args.nemsis_year, args.val_file_scene)

    # Initialize variables
    rows = []
    columns = None

    # Read the CSV file
    with open(test_file_scene_path, 'r') as file:
        # Read the file line by line and split by the custom delimiter '~|~'
        lines = file.readlines()
        columns = lines[0].strip().split('~|~')  # Get header
        num_columns = len(columns)
        # columns = [col.strip().replace('|', '').replace(',', '') for col in columns]  # Clean headers

        for line in lines[1:]:  # Skip the header row
            row = line.strip().split('~|~')  # Split each row by the custom delimiter
            assert len(row) == num_columns, f"Row has {len(row)} columns, expected {num_columns}"
            # row = [cell.strip().replace('|', '').replace(',', '') for cell in row]
            rows.append(row)

    # Create a list of dicts where each dict represents a row
    data = [dict(zip(columns, row)) for row in rows]

    # 1. Filter out rows where "med_type" has more than one value
    filtered_data = [row for row in data if (len(row['med_type'].split(' ')) == 1 and len(row['protocol'].split(' ')) == 1)]

    # Sample 45+5 rows using util.global_seed
    random.seed(util.global_seed)
    sampled_data = random.sample(filtered_data, 46)

    # Add the rows with the specified PCR keys
    pcr_rows_to_add = [row for row in data if row['pcr'] in pcr_keys_to_add]

    # Combine and ensure there are no duplicates based on PCR key
    combined_data = {row['pcr']: row for row in pcr_rows_to_add + sampled_data}
    sampled_data = list(combined_data.values())  # Convert back to list

    # Ensure we have exactly 50 examples
    if len(sampled_data) > 50:
        sampled_data = sampled_data[:51]

    # 2. Translate the "ps", "pi", "as", "si", and "protocol" columns
    for row in sampled_data:
        for col in ['ps', 'pi', 'as', 'si']:
            si_desc_list = []
            for si_code in row[col].split(' '):
                si_desc_list.append(my_np.global_d[si_code])
            si_desc = ' '.join(si_desc_list)
            row[col] = si_desc
        # assert len(row['protocol'].split()) == 1, f"Row has more than one protocol: {row['protocol']}"
        protocol_id = my_np.dedicated_protocols[int(row['protocol'])]
        row['protocol'] = my_np.nemsis_protocol_id2desc[protocol_id]

    # 3. Replace "nos" with "not otherwise specified" in the specified columns
    for row in sampled_data:
        for col in ['ps', 'pi', 'as', 'si']:
            # row[col] = row[col].replace("nos", "not otherwise specified")
            row[col] = re.sub(r'\bnos\b', 'not otherwise specified', row[col])

    # 4. Replace "scene" column with "pill exist" and "alcohol exist"
    for row in sampled_data:
        row['alcohol exist'] = 'yes' if '1' in row['scene'] else 'no'
        row['pill exist'] = 'yes' if '2' in row['scene'] else 'no'
        del row['scene']  # Remove the original "scene" column

    # 5. Rename vital columns
    vital_mapping = {
        'vital_0': 'blood pressure',
        'vital_1': 'heart rate',
        'vital_2': 'pulse oximetry',
        'vital_3': 'respiratory rate',
        'vital_4': 'etco2',
        'vital_5': 'blood glucose',
        'ps': 'primary symptom',
        'pi': 'primary impression',
        'as': 'associate symptom',
        'si': 'secondary impression'        
    }

    for row in sampled_data:
        for old_col, new_col in vital_mapping.items():
            row[new_col] = row.pop(old_col)  # Rename each column

    # 6. Add "inferred disease" and "dosage needed" columns
    for row in sampled_data:
        row['inferred disease'] = ''
        row['dosage needed'] = ''

    # 7. Write the result to a CSV file
    # user_study_file_path = os.path.join(cache_dir, args.nemsis_year, "user_study_50.csv")


    # Manually write the CSV file with the custom delimiter
    with open(user_study_file_path, 'w') as file:
        # Write the header
        file.write('~|~'.join(sampled_data[0].keys()) + '\n')
        # Write each row
        for row in sampled_data:
            file.write('~|~'.join(str(value) for value in row.values()) + '\n')

    print(f"50 samples written to {user_study_file_path}")

def tokenize_function(text_encoder_name, text, max_seq_len):
    tokenizer = AutoTokenizer.from_pretrained(text_encoder_name)
    return tokenizer(text, padding="max_length", truncation=True,  max_length=max_seq_len)

def predict_emsglass_single_sample(example, 
                                   args,
                                   text_model=None,
                                   vital_model=None,
                                   scene_model=None, 
                                   text_vital_model=None,
                                   text_scene_model=None, 
                                   vital_scene_model=None,
                                   text_vital_scene_model=None):
    # Map models to required inputs
    model_input_map = {
        "text_model": (text_model, ["input_ids", "attention_mask", "token_type_ids"]),
        "vital_model": (vital_model, ["vital"]),
        "scene_model": (scene_model, ["scene"]),
        "text_vital_model": (text_vital_model, ["input_ids", "attention_mask", "token_type_ids", "vital"]),
        "text_scene_model": (text_scene_model, ["input_ids", "attention_mask", "token_type_ids", "scene"]),
        "vital_scene_model": (vital_scene_model, ["vital", "scene"]),
        "text_vital_scene_model": (text_vital_scene_model, ["input_ids", "attention_mask", "token_type_ids", "vital", "scene"])
    }

    device = torch.device(f"cuda:{args.device}" if torch.cuda.is_available() else "cpu")
    # sample = {k: v.to(device) for k, v in sample.items()}  # Move sample data to device
    text, vital, scene = example
    tokenized_text_ids = tokenize_function(args.text_encoder_name, text, args.max_seq_len)
    sample = {k: torch.tensor(v).to(device) for k, v in tokenized_text_ids.items()}
    sample.update({"vital": vital.to(device), "scene": scene.to(device)})
    sample = {k: v.unsqueeze(0) for k, v in sample.items()}  # Add batch dimension
    # print("predict sample: ", sample)
    # print("predict sample input_ids shape: ", sample["input_ids"].shape)
    # print("predict sample vital shape: ", sample["vital"].shape)
    # print("predict sample scene shape: ", sample["scene"].shape)
    
    all_results = {}

    for model_name, (model, required_inputs) in model_input_map.items():
        if model is None:
            continue
        print(f"\nPredicting with {model_name}...")
        model.to(device)
        model.eval()

        # Prepare inputs for the model
        inputs = {k: sample[k] for k in required_inputs if k in sample}
        
        with torch.no_grad():
            # Run prediction
            protocol_logits, med_type_logits, quantity_output, *_ = model(**inputs)

        # Process and save predictions
        predictions = {
            "protocol": torch.argmax(protocol_logits, dim=-1),
            "med_type": torch.argmax(med_type_logits, dim=-1),
            "quantity": quantity_output.squeeze(-1)
        }

        # Store predictions for each model
        all_results[model_name] = predictions

    return all_results

# Function to load specific samples from the CSV
def load_samples_from_csv(args, csv_path, indices, vital_k):

    df = pd.read_csv(csv_path, delimiter='~|~', engine='python')
    text_data = []
    vital_data = []
    scene_data = []

    for idx in indices:
        print(f"\nSample {idx + 1}:")
        sample = df.iloc[idx]

        # text_data
        concatenated_symptom = sample["primary symptom"] + " " + \
                                sample["primary impression"] + " " + \
                                sample["associate symptom"] + " " + \
                                sample["secondary impression"]
        transformed_symptom_list = util.text_transformation(concatenated_symptom)[0]
        text = " ".join(transformed_symptom_list)
        print(text)
        text_data.append(text)

        # sequence_length x vitals_num
        vitals = np.array([str(sample[k]).split(' ') for k in vital_k], dtype=np.float32).T
        # vitals = torch.from_numpy(vitals, dtype=torch.float32)
        print(vitals)
        vital_data.append(torch.from_numpy(vitals))

        one_hot_scene = torch.zeros(args.scene_input_size, dtype=torch.float32)
        if sample["alcohol exist"] == "yes":
            one_hot_scene[0] = 1
        elif sample["pill exist"] == "yes":
            one_hot_scene[1] = 1
        else:
            print("No scene")
        print(one_hot_scene)
        scene_data.append(one_hot_scene)
        
    return zip(text_data, vital_data, scene_data)

def predict_emsglass(args,
                     vital_k,
                    text_model=None,
                    vital_model=None,
                    scene_model=None, 
                    text_vital_model=None,
                    text_scene_model=None, 
                    vital_scene_model=None,
                    text_vital_scene_model=None):
    
    csv_path = "/home/liuyi/transformers/data/nemsis_cache_files/2023/user_study_50.csv"
    target_indices = [0, 2]  # 1st and 3rd sample (0-based index)
    
    # Load samples and perform predictions
    samples = load_samples_from_csv(args, csv_path, target_indices, vital_k)
    # print("Loaded samples:", samples)
    # return 
    all_results = []
    
    for sample in samples:
        results = predict_emsglass_single_sample(
            sample, args, text_model, vital_model, scene_model, text_vital_model, text_scene_model, vital_scene_model, text_vital_scene_model
        )
        all_results.append(results)
        print("Results for sample:", results)
    
    return all_results


# if __name__ == "__main__":
    
#   time_s = datetime.now()

#   parser = argparse.ArgumentParser(description = "control the functions for NEMSIS Medication processing")
#   parser.add_argument("--home_dir", action='store', type=str, default = "/home/liuyi/transformers")
#   parser.add_argument("--nemsis_dir", action='store', type=str, default = "/slot1/NEMSIS_Databases")
#   parser.add_argument("--nemsis_year", action='store', type=str, default = "2021")
#   parser.add_argument("--data_folder", action='store', type=str, default = "data")
#   parser.add_argument("--cache_folder", action='store', type=str, default = "nemsis_cache_files")
#   parser.add_argument("--dedicated_med_file", action='store', type=str, default = "tamu_medication.csv")
#   parser.add_argument("--nemsis_med_quant_unit_file", action='store', type=str, default = "nemsis_med_quant_units.csv")
#   parser.add_argument("--device", action='store', type=str, default = "0")
#   parser.add_argument("--saved_model_dir", action='store', type=str, default = "/slot1/ems_models/EMSFoudation_models/")
#   parser.add_argument("--patience", action='store', type=int, default = 3)
#   parser.add_argument("--epoch", action='store', type=int, default = 20)
#   parser.add_argument("--debug_mode", action='store_true', help="Prefixing the model names with debug_mode")

#   parser.add_argument("--max_seq_len", action='store', type=int, default=128)
#   parser.add_argument("--input_size", action='store', type=int, default=6)
#   parser.add_argument("--hidden_size", action='store', type=int, default=64)
#   parser.add_argument("--num_layers", action='store', type=int, default=1)  
#   parser.add_argument("--dropout_prob", action='store', type=float, default=0.0)
#   parser.add_argument("--pre_padding", action='store_true', help="Pre-padding the vitals")
#   parser.add_argument("--padding_value", action='store', type=float, default=0.0)
#   parser.add_argument("--scene_input_size", action='store', type=int, default=2)

#   parser.add_argument("--text_encoder_name", type=str, default = None, choices=[
#       'bert-base-cased', 'bert-base-uncased', 'bert-large-cased', 'bert-large-uncased','arnabdhar/tinybert-imdb', 'google/mobilebert-uncased'])
#   parser.add_argument("--vital_encoder_name", type=str, default = None, choices=['rnn', 'gru', 'lstm'])
#   parser.add_argument("--scene_encoder_name", type=str, default = None, choices=['fc'])
#   parser.add_argument("--vitals", type=str, default="012345")
#   parser.add_argument("--fill", type=str, default="mean", choices=['mean', 'forward_backward'])
#   parser.add_argument("--norm", type=str, required=True, choices=['minmax', 'z_score', 'chained'])

#   parser.add_argument("--train_batch_size", action='store', type=int, default = 8)
#   parser.add_argument("--validation_batch_size", action='store', type=int, default = 32)
#   parser.add_argument("--test_batch_size", action='store', type=int, default = 32)

# #   parser.add_argument("--hg_config_name", type=str, default="si_protocol_med_vitals_history", choices=['med_labels', 'si_protocol_med_vitals_history', 'scene'])
#   parser.add_argument("--hg_config_name", type=str, default="scene", choices=['med_labels', 'si_protocol_med_vitals_history', 'scene'])
#   parser.add_argument("--train_file", action='store', type=str, default = "train_file_noscene.txt")
#   parser.add_argument("--val_file", action='store', type=str, default = "val_file_noscene.txt")
#   parser.add_argument("--test_file", action='store', type=str, default = "test_file_noscene.txt")
#   parser.add_argument("--train_file_scene", action='store', type=str, default = "train_file_scene.txt")
#   parser.add_argument("--val_file_scene", action='store', type=str, default = "val_file_scene.txt")
#   parser.add_argument("--test_file_scene", action='store', type=str, default = "test_file_scene.txt")

#   parser.add_argument("--single_label", action='store_true', default=True, help="Perform single-label classifications")
#   parser.add_argument("--do_protocol", action='store_true', help="Perform protocol classification")
#   parser.add_argument("--do_med_type", action='store_true', help="Perform medicine type classification")
#   parser.add_argument("--do_quantity", action='store_true', help="Perform quantity regression")



#   args = parser.parse_args()
#   set_seed(util.global_seed)

#   run_train(args)

#   time_t = datetime.now() - time_s
#   print("This run takes %s" % time_t)